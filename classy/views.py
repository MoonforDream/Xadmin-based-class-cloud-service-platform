import json
import os
import re
import shutil
import urllib
import zipfile
import pandas as pd
import redis
from django.contrib.auth.hashers import make_password
from django.db.models import IntegerField
from django.db.models.functions import Cast
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from pypinyin import lazy_pinyin
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.utils.encoding import force_bytes
from geopy.distance import distance
from django.http import JsonResponse, HttpResponse, FileResponse
import requests
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from mycelery.email.tasks import send_sms, send_pass, send_email, send_repass
from classy.models import task, UserInfo
import uuid
import time
from django.core.paginator import Paginator, PageNotAnInteger


# Create your views here.
@login_required
def index(request):
    if request.method=='GET':
        login_user=request.user
        user=UserInfo.objects.get(username=login_user)
        tasks=task.objects.filter(clss_id=user.class_num,task_status='进行中')
        is_ti=0
        if tasks:
            for i in tasks:
                feedback = i.task_feedback.split(',')
                if user.stu_name in feedback:
                    is_ti=1
        return render(request,'index.html',locals())

def login_view(request):
    if request.method=='GET':
        msg=request.session.get('message')
        if msg:
            message=msg
            del request.session['message']
        else:
            message=''
        msg1=request.session.get('message1')
        if msg1:
            message1=msg1
            del request.session['message1']
        else:
            message1=''
        return render(request,'login.html',locals())
    if request.method=='POST':
        username=request.POST.get('username')
        password=request.POST.get('password')
        user=authenticate(request,username=username,password=password,is_active=1)
        if user:
            login(request,user)
            return redirect('/class/admin/')
        else:
            error='该用户不存在或者帐号密码错误'
            return render(request,'login.html',{'error':error})




@login_required
def task_list(request):
    login_user = request.user
    user = UserInfo.objects.get(username=login_user)
    if request.method=='GET':
        task_lst=[]
        class_id=user.class_num
        tasks=task.objects.filter(clss_id=class_id).order_by('-id')
        conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=7)
        send=user.stu_name
        for i in tasks:
            feedback_lst=i.task_feedback.split(',')
            if not feedback_lst or feedback_lst==['']:
                i.task_status='已完成'
                i.save()
            is_status=conn.get(i.task_id)
            if user.stu_name not in feedback_lst and user.stu_name!= i.send_name:
                i.task_status='已完成'
                tk = {'task_id': i.task_id, 'task_name': i.task_name, 'task_type': i.task_type,
                      'task_start': i.task_start, 'task_end': i.task_end, 'task_status': i.task_status,'send_name':i.send_name}
            elif is_status:
                tk={'task_id':i.task_id,'task_name':i.task_name,'task_type':i.task_type,'task_start':i.task_start,'task_end':i.task_end,'task_status':i.task_status,'send_name':i.send_name}
            else:
                i.task_status='已截止'
                tk = {'task_id': i.task_id, 'task_name': i.task_name, 'task_type': i.task_type,
                      'task_start': i.task_start, 'task_end': i.task_end, 'task_status': i.task_status,'send_name':i.send_name}
                i.save()
            task_lst.append(tk)
        p=Paginator(task_lst,7)
        try:
            page_number = request.GET.get('page', 1)
        except PageNotAnInteger:
            page_number = 1
        page_obj = p.page(page_number)
        return render(request,'task-list.html', locals())
    if request.method=='POST':
        print(request.POST)
        print(request)
        if 'start' in request.POST:
            task_lst=[]
            start=request.POST.get('start')
            end=request.POST.get('end')
            taskname=request.POST.get('username')
            lst=''
            if start and not end and not taskname:
                lst=task.objects.filter(task_start__contains=start,clss_id=user.class_num)
            elif not start and end and not taskname:
                lst=task.objects.filter(task_end__contains=end,clss_id=user.class_num)
            elif not start and not end and taskname:
                lst=task.objects.filter(task_name__contains=taskname,clss_id=user.class_num)
            elif start and end and taskname:
                lst=task.objects.filter(task_name__contains=taskname,task_start__contains=start,task_end__contains=end,clss_id=user.class_num)
            elif start and end and not taskname:
                lst=task.objects.filter(task_start__contains=start,task_end__contains=end,clss_id=user.class_num)
            elif start and not end and taskname:
                lst=task.objects.filter(task_start__contains=start,task_name__contains=taskname,clss_id=user.class_num)
            elif not start and end and taskname:
                lst=task.objects.filter(task_name__contains=taskname,task_end__contains=end,clss_id=user.class_num)
            if not lst:
                m1='未找到该任务'
            conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=7)
            send = user.stu_name
            for i in lst:
                feedback_lst = i.task_feedback.split(',')
                if not feedback_lst:
                    i.task_status = '已完成'
                    i.save()
                is_status = conn.get(i.task_id)
                if user.stu_name not in feedback_lst and user.stu_name != i.send_name:
                    i.task_status = '已完成'
                    tk = {'task_id': i.task_id, 'task_name': i.task_name, 'task_type': i.task_type,
                          'task_start': i.task_start, 'task_end': i.task_end, 'task_status': i.task_status,
                          'send_name': i.send_name}
                elif is_status:
                    tk = {'task_id': i.task_id, 'task_name': i.task_name, 'task_type': i.task_type,
                          'task_start': i.task_start, 'task_end': i.task_end, 'task_status': i.task_status,
                          'send_name': i.send_name}
                else:
                    i.task_status = '已截止'
                    tk = {'task_id': i.task_id, 'task_name': i.task_name, 'task_type': i.task_type,
                          'task_start': i.task_start, 'task_end': i.task_end, 'task_status': i.task_status,
                          'send_name': i.send_name}
                    i.save()
                task_lst.append(tk)
            p = Paginator(task_lst, 7)
            try:
                page_number = request.GET.get('page', 1)
            except PageNotAnInteger:
                page_number = 1
            page_obj = p.page(page_number)
            return render(request,'task-list.html',locals())
        task_id = request.POST.get('task_id')
        if not task_id:
            task_id=request.session.get('task_id')
            del request.session['task_id']
        obj = task.objects.get(task_id=task_id)
        task_status = obj.task_status
        if task_status=='已截止':
            message='任务已截止'
            return HttpResponse(message)
        elif task_status=='已完成':
            message='任务已经完成'
            return HttpResponse(message)
        if obj.task_type=='签到':
            radius=obj.radius
            checkin_coordinate = (obj.lat, obj.lon)
            url = 'http://txt.go.sohu.com/ip/soip'
            response = requests.get(url)
            text = response.text
            ip = re.findall(r'\d+.\d+.\d+.\d+', text)[0]
            r1 = requests.get("http://ip-api.com/json/%s" %ip).text
            json2 = json.loads(r1)
            user_coordinate = (json2['lat'], json2['lon'])
            distance1=distance(checkin_coordinate,user_coordinate)
            if distance1 <= radius:
                feedback_lst=obj.task_feedback.split(',')
                if user.stu_name in feedback_lst:
                    feedback_lst.remove(user.stu_name)
                feedback_lst=",".join(feedback_lst)
                obj.task_feedback=feedback_lst
                obj.save()
                message='签到成功'
            else:
                message = "不在签到范围内"
            return HttpResponse(message)
        if obj.task_type=='作业':
            files = request.FILES['file']
            file_path='/home/dreamcreator/class/media/%s/%s'%(str(task_id),files.name)
            with open(file_path, 'wb') as file:
                for i in files.chunks():
                    file.write(i)
            feedback_lst = obj.task_feedback.split(',')
            if user.stu_name in feedback_lst:
                feedback_lst.remove(user.stu_name)
            feedback_lst = ",".join(feedback_lst)
            obj.task_feedback = feedback_lst
            obj.save()
            return JsonResponse({'file':file_path})

@login_required
def queue(request):
    login_user = request.user
    obj = UserInfo.objects.get(username=login_user)
    if request.method=='GET':
        task_id = request.GET.get('task_id')
        request.session['id']=task_id
        tasks = task.objects.get(task_id=task_id, clss_id=obj.class_num)
        size = int(tasks.describe)-1
        return render(request,'queue.html',locals())
    if request.method=='POST':
        task_id=request.session.get('id')
        if task_id:
            del request.session['id']
        tasks=task.objects.get(task_id=task_id,clss_id=obj.class_num)
        size=tasks.describe
        leader=request.POST.get('leader')
        member=[]
        wb = load_workbook('/home/dreamcreator/class/media/xlsx/%s.xlsx' % task_id)
        ws = wb.active
        for i in range(1,int(size)+1):
            member1='member-%s'%i
            if member1 in request.POST:
                member.append(request.POST.get(member1))
        feedback_lst = tasks.task_feedback.split(',')
        if leader in feedback_lst:
            feedback_lst.remove(leader)
        for j in member:
            if j in feedback_lst:
                feedback_lst.remove(j)
        member=",".join(member)
        ws.append([leader,member])
        wb.save('/home/dreamcreator/class/media/xlsx/%s.xlsx' % task_id)
        feedback_lst = ",".join(feedback_lst)
        tasks.task_feedback=feedback_lst
        tasks.save()
        return render(request,'queue.html')


@login_required
def role(request):
    login_user = request.user
    obj = UserInfo.objects.get(username=login_user)
    if request.method=='GET':
        if obj.is_staff==0 or obj.stu_work not in ['学习委员','学委','班长','副班长']:
            m='你没有权限使用该功能'
            request.session['m']=m
            return redirect('/class/welcome/')
        user=UserInfo.objects.filter(class_num=obj.class_num).order_by('-stunum').annotate(
    stunum_int=Cast('stunum', IntegerField())
).order_by('stunum_int')
        stu_list=[]
        for i in user:
            if i.email=='':
                i.email='无'
                i.save()
            if i.is_active==1:
                i.is_active='已启用'
            else:
                i.is_active='已停用'
            tk={'stunum':i.stunum,'stu_id':i.stu_id,'username':i.username,'stu_name':i.stu_name,'email':i.email,'stu_work':i.stu_work,'stu_status':i.is_active}
            stu_list.append(tk)
        p=Paginator(stu_list,7)
        try:
            page_number = request.GET.get('page', 1)
        except PageNotAnInteger:
            page_number = 1
        page_obj = p.page(page_number)
        return render(request,'admin-role.html',locals())
    if request.method=='POST':
        if 'stu' in request.POST:
            stu = request.POST.get('stu')
            username = request.POST.get('username')
            user1=''
            if not stu and username:
                user1 = UserInfo.objects.filter(stu_name__contains=username, class_num=obj.class_num)
            elif not username and stu:
                user1 = UserInfo.objects.filter(stu_id__exact=stu, class_num=obj.class_num)
            else:
                user1 = UserInfo.objects.filter(stu_id__exact=stu, stu_name__contains=username, class_num=obj.class_num)
            if not user1:
                message = '未找到该学生'
            stu_list = []
            for i in user1:
                if i.email == '':
                    i.email = '无'
                    i.save()
                if i.is_active == 1:
                    i.is_active = '已启用'
                else:
                    i.is_active = '已停用'
                tk = {'stunum': i.stunum, 'stu_id': i.stu_id, 'username': i.username, 'stu_name': i.stu_name,
                      'email': i.email, 'stu_work': i.stu_work, 'stu_status': i.is_active}
                stu_list.append(tk)
            p = Paginator(stu_list, 7)
            try:
                page_number = request.GET.get('page', 1)
            except PageNotAnInteger:
                page_number = 1
            page_obj = p.page(page_number)
            return render(request, 'admin-role.html', locals())
        try:
            stu_id=request.POST.get('studentId')
            dele=UserInfo.objects.get(username=stu_id,class_num=obj.class_num)
            dele.delete()
        except:
            student_ids = request.POST.getlist('student_ids[]')
            for student_id in student_ids:
                dele=UserInfo.objects.get(stu_id=student_id,class_num=obj.class_num)
                dele.delete()
        return JsonResponse({'status': 'success', 'message': '删除成功！'})

@login_required
def changestatus(request):
    if request.method == 'POST':
        username=request.POST.get('studentId')
        obj = UserInfo.objects.get(username=username)
        if obj.is_active == 1:
            obj.is_active = 0
            obj.save()
            return JsonResponse({'status': 'success', 'message': '停用成功！'})
        elif obj.is_active == 0:
            obj.is_active = 1
            obj.save()
            return JsonResponse({'status': 'success', 'message': '启用成功！'})


@login_required
def form(request):
    if request.method=='GET':
        login_user=request.user
        obj=UserInfo.objects.get(username=login_user)
        if obj.is_staff==0:
            m='你没有权限使用该功能'
            request.session['m']=m
            return redirect('/class/welcome/')
        return render(request,'form.html')
    if request.method=='POST':
        login_user=request.user
        user=UserInfo.objects.get(username=login_user)
        class_id=user.class_num
        if 'radius' in request.POST:
            cnt=0
            end=request.POST.get('end')
            radius=request.POST.get('radius')
            if end and radius:
                url = 'http://txt.go.sohu.com/ip/soip'
                response = requests.get(url)
                text = response.text
                ip = re.findall(r'\d+.\d+.\d+.\d+', text)[0]
                r1=requests.get("http://ip-api.com/json/%s"%ip).text
                json2=json.loads(r1)
                task_id=uuid.uuid4()
                stu=UserInfo.objects.filter(class_num=class_id)
                stu_lst=''
                for i in stu:
                    if user.stu_name==i.stu_name:
                        continue
                    if cnt==0:
                        stu_lst+=str(i.stu_name)
                    else:
                        stu_lst+=','+str(i.stu_name)
                    cnt+=1
                obj1=task(task_id=task_id,task_end=end,lat=json2['lat'],lon=json2['lon'],clss_id=class_id,task_type='签到',radius=radius,task_feedback=stu_lst,send_name=user.stu_name)
                obj1.save()
                tasks=task.objects.get(task_id=task_id)
                start=tasks.task_start
                start = time.struct_time(start.timetuple())
                start=time.mktime(start)
                end=time.mktime(time.strptime(end,"%Y-%m-%d %H:%M:%S"))
                diff=int(end-start)
                if diff>=0:
                    conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=7)
                    task_id=str(task_id)
                    conn.set(task_id, 1, ex=diff)
                    message='发布成功'
                else:
                    message1='无法发布已经过期的任务'
                    tasks.delete()
        if 'work' in request.POST:
            files = request.FILES.getlist('file')
            print(files)
            cnt=0
            work=request.POST.get('work')
            end1=request.POST.get('end1')
            describe=request.POST.get('describe')
            task_id=request.session.get('task')
            if not task_id:
                task_id = uuid.uuid4()
            else:
                del request.session['task']
            stu = UserInfo.objects.filter(class_num=class_id)
            stu_lst = ''
            path=request.session.get('task_path')
            for i in stu:
                if user.stu_name == i.stu_name:
                    continue
                if cnt == 0:
                    stu_lst += str(i.stu_name)
                else:
                    stu_lst += ',' + str(i.stu_name)
                cnt += 1
            obj = task(task_id=task_id, task_name=work, task_type='作业', task_end=end1, clss_id=class_id,
                       task_feedback=stu_lst, describe=describe, send_name=user.stu_name)
            if path:
                del request.session['task_path']
                obj.task_file=path
            obj.save()
            tasks=task.objects.get(task_id=task_id)
            start = tasks.task_start
            start = time.struct_time(start.timetuple())
            start = time.mktime(start)
            end = time.mktime(time.strptime(end1, "%Y-%m-%d %H:%M:%S"))
            diff = int(end - start)
            if diff >= 0:
                conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=7)
                task_id = str(task_id)
                conn.set(task_id, 1, ex=diff)
                message = '发布成功'
                os.mkdir('/home/dreamcreator/class/media/%s' % str(task_id))
            else:
                message1 = '无法发布已经过期的任务'
                tasks.delete()
        if 'size' in request.POST:
            task_name=request.POST.get('task')
            end2=request.POST.get('end2')
            size=request.POST.get('size')
            cnt=0
            task_id=uuid.uuid4()
            stu = UserInfo.objects.filter(class_num=class_id)
            stu_lst = ''
            for i in stu:
                if user.stu_name == i.stu_name:
                    continue
                if cnt == 0:
                    stu_lst += str(i.stu_name)
                else:
                    stu_lst += ',' + str(i.stu_name)
                cnt += 1
            obj=task(task_name=task_name,task_end=end2,describe=size,task_id=task_id,task_type='分组任务',send_name=user.stu_name,task_feedback=stu_lst,clss_id=user.class_num)
            obj.save()
            tasks = task.objects.get(task_id=task_id)
            start = tasks.task_start
            start = time.struct_time(start.timetuple())
            start = time.mktime(start)
            end = time.mktime(time.strptime(end2, "%Y-%m-%d %H:%M:%S"))
            diff = int(end - start)
            if diff >= 0:
                conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=7)
                task_id = str(task_id)
                conn.set(task_id, 1, ex=diff)
                message = '发布成功'
                wb=Workbook()
                ws=wb.active
                ws.append([tasks.task_name])
                ws.append(['组长：','组员：'])
                wb.save('/home/dreamcreator/class/media/xlsx/%s.xlsx'%tasks.task_id)
            else:
                message1 = '无法发布已经过期的任务'
                tasks.delete()
        return render(request, 'form.html',locals())



def getfile(request):
    if request.method == 'POST':
        files = request.FILES.getlist('file')
        task_id = uuid.uuid4()
        request.session['task'] = str(task_id)
        path = '/home/dreamcreator/class/media/uploads/' + str(task_id) + 'upload'
        os.makedirs(path, exist_ok=True)
        urls = []
        for f in files:
            file_path='%s/%s'%(path,f.name)
            with open(file_path,'wb') as file:
                for i in f.chunks():
                    file.write(i)
            urls.append(file_path)
        request.session['task_path']=file_path
        return JsonResponse({'code': 0, 'msg': '上传成功', 'data': {'urls': urls,'id':123}})
    else:
        return render(request, 'form.html')

@login_required
def upload(request):
    login_user=request.user
    obj=UserInfo.objects.get(username=login_user)
    class_id=obj.class_num
    stu_class=obj.stu_class
    if request.method=='POST':
        files = request.FILES['file']
        df=pd.read_excel(files,header=2)
        num_rows = len(df)
        data=[]
        for i in range(num_rows):
            data.append(df.values[i])
            uidb64 = urlsafe_base64_encode(force_bytes(df.values[i][0]))
            pinyin = lazy_pinyin(df.values[i][2])
            pinyin_abbr = ''.join([s[0] for s in pinyin])
            username=str(uidb64)+str(df.values[i][1])+pinyin_abbr
            num=str(df.values[i][1])[-4:]  #学号后四位
            password='ZKSD'+pinyin_abbr+num
            obj1=UserInfo.objects.filter(stu_name=df.values[i][2])
            if obj1:
                continue
            user = UserInfo.objects.create_user(username=username, password=password, is_active=1,is_superuser=0,class_num=class_id,stu_name=df.values[i][2],stu_id=df.values[i][1],stunum=df.values[i][0],stu_class=stu_class,is_staff=0)
            user.save()
        return JsonResponse({'file': '/class/upload/' + files.name})
    else:
        return render(request,'stu_upload.html')

def password(request):
    if request.method=='GET':
        return render(request,'member-password.html')
    if request.method=='POST':
        username=request.POST.get('username')
        email=request.POST.get('email')
        password=request.POST.get('password')
        user=UserInfo.objects.get(username=username)
        if email==user.email and not user.check_password(password):
            send_repass(email,user.id)
            request.session['changepass']=password
            message1 = '密码修改激活链接已通过邮箱发送成功，请尽快激活！'
        elif user.check_password(password):
            message='新密码与原密码相同，无法更改!'
        else:
            message='邮箱与帐号邮箱不符！'
        return render(request,'member-password.html',locals())

def register(request):
    if request.method=='POST':
        username=request.POST.get('username')
        password=request.POST.get('password')
        email=request.POST.get('email')
        obj1=UserInfo.objects.filter(email=email)
        if obj1:
            message = '该邮箱注册过了'
        obj = UserInfo.objects.filter(username=username)
        if obj:
            message = '该用户名已经注册过了'
        if not obj and not obj1:
            user=UserInfo.objects.create_user(username=username,password=password,email=email,is_active=0,is_superuser=0,is_staff=1)
            obj1=UserInfo.objects.get(username=username)
            uid=obj1.id
            result = send_sms.delay(email, uid)
            message1='成功注册，邮箱激活后即可使用'
        return render(request,'member-edit.html',locals())
    else:
        return render(request,'member-edit.html')

@login_required
def committee(request):
    login_user=request.user
    obj = UserInfo.objects.get(username=login_user)
    if request.method=='GET':
        user=UserInfo.objects.filter(class_num=obj.class_num)
        stu_list=[]
        for i in user:
            if i.stu_work!='学生':
                if i.is_active==1:
                    i.is_active='已启用'
                else:
                    i.is_active='已停用'
                tk={'stunum':i.stunum,'stu_id':i.stu_id,'username':i.username,'stu_name':i.stu_name,'email':i.email,'stu_work':i.stu_work,'stu_status':i.is_active}
                stu_list.append(tk)
        p = Paginator(stu_list, 7)
        try:
            page_number = request.GET.get('page', 1)
        except PageNotAnInteger:
            page_number = 1
        page_obj = p.page(page_number)
        return render(request,'class-role.html',locals())
    if request.method=='POST':
        if 'stu' or 'username' in request.POST:
            stu = request.POST.get('stu')
            username = request.POST.get('username')
            user1=''
            if not stu and username:
                user1 = UserInfo.objects.filter(stu_name__contains=username, class_num=obj.class_num)
            elif not username and stu:
                user1 = UserInfo.objects.filter(stu_id__exact=stu, class_num=obj.class_num)
            else:
                user1 = UserInfo.objects.filter(stu_id__exact=stu, stu_name__contains=username, class_num=obj.class_num)
            if not user1:
                message = '未找到该学生'
            stu_list = []
            for i in user1:
                if i.email == '':
                    i.email = '无'
                    i.save()
                if i.is_active == 1:
                    i.is_active = '已启用'
                else:
                    i.is_active = '已停用'
                tk = {'stunum': i.stunum, 'stu_id': i.stu_id, 'username': i.username, 'stu_name': i.stu_name,
                      'email': i.email, 'stu_work': i.stu_work, 'stu_status': i.is_active}
                stu_list.append(tk)
            p = Paginator(stu_list, 7)
            try:
                page_number = request.GET.get('page', 1)
            except PageNotAnInteger:
                page_number = 1
            page_obj = p.page(page_number)
            return render(request, 'class-role.html', locals())


@login_required
def help(request):
    return render(request,'help.html')

@login_required
def welcome(request):
    cnt=0   #发布作业数
    ans=0   #提交作业数
    n=0     #发布签到数
    ns=0    #完成签到数
    class_num=0    #班级人数
    login_user=request.user
    user=UserInfo.objects.get(username=login_user)
    user_list=UserInfo.objects.filter(class_num=user.class_num)
    tasks=task.objects.filter(clss_id=user.class_num)
    num=tasks.count
    for i in tasks:
        if i.send_name==user.stu_name and i.task_type!='签到':
            cnt+=1
        if i.send_name==user.stu_name and i.task_type=='签到':
            n+=1
        feedback=i.task_feedback.split(',')
        if i.send_name!=user.stu_name and user.stu_name in feedback and i.task_type!='签到':
            ans+=1
        if i.send_name != user.stu_name and user.stu_name in feedback and i.task_type == '签到':
            ns+=1
    for j in user_list:
        if j:
            class_num+=1
    m = request.session.get('m')
    if m:
        del request.session['m']
    return render(request,'welcome.html',locals())

@login_required
def log(request):
    return render(request,'log.html')

@login_required
def feedback(request):
    login_user = request.user
    obj = UserInfo.objects.get(username=login_user)
    if request.method=='GET':
        task_lst=[]
        class_id=obj.class_num
        task_list=task.objects.filter(clss_id=class_id).order_by('-id')
        for i in task_list:
            tk={'task_id':i.task_id,'task_name':i.task_name,'task_type':i.task_type,'start':i.task_start,'end':i.task_end,'feedback':i.task_feedback}
            task_lst.append(tk)
        p = Paginator(task_lst, 7)
        try:
            page_number = request.GET.get('page', 1)
        except PageNotAnInteger:
            page_number = 1
        page_obj = p.page(page_number)
    if request.method=='POST':
        if 'start' or 'end' or 'username' in request.POST:
            start=request.POST.get('start')
            end=request.POST.get('end')
            taskname=request.POST.get('username')
            task_lst=[]
            lst = ''
            if start and not end and not taskname:
                lst = task.objects.filter(task_start__contains=start, clss_id=obj.class_num)
            elif not start and end and not taskname:
                lst = task.objects.filter(task_end__contains=end, clss_id=obj.class_num)
            elif not start and not end and taskname:
                lst = task.objects.filter(task_name__contains=taskname, clss_id=obj.class_num)
            elif start and end and taskname:
                lst = task.objects.filter(task_name__contains=taskname, task_start__contains=start,
                                          task_end__contains=end, clss_id=obj.class_num)
            elif start and end and not taskname:
                lst=task.objects.filter(task_start__contains=start,task_end__contains=end,clss_id=obj.class_num)
            elif start and not end and taskname:
                lst=task.objects.filter(task_start__contains=start,task_name__contains=taskname,clss_id=obj.class_num)
            elif end and not start and taskname:
                lst=task.objects.filter(task_end__contains=end,task_name__contains=taskname,clss_id=obj.class_num)
            if not lst:
                m1 = '未找到该任务'
            conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=7)
            send = obj.stu_name
            for i in lst:
                tk = {'task_id': i.task_id, 'task_name': i.task_name, 'task_type': i.task_type, 'start': i.task_start,
                      'end': i.task_end, 'feedback': i.task_feedback}
                task_lst.append(tk)
            p = Paginator(task_lst, 7)
            try:
                page_number = request.GET.get('page', 1)
            except PageNotAnInteger:
                page_number = 1
            page_obj = p.page(page_number)
    return render(request,'feedback.html',locals())



def activate(request):
    uidb64=request.GET.get('uid')
    token=request.GET.get('token')
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        # 获取对应的用户对象
        user =UserInfo.objects.get(id=uid)
    except (TypeError, ValueError, OverflowError, UserInfo.DoesNotExist):
        user = None
    conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=3)
    is_token = conn.get(user.id).decode('utf-8')
    if is_token:
        if user is not None and token==is_token:
            user.is_active=True
            snowflake = Snowflake(1)
            user.class_num=snowflake.generate_id()
            user.save()
            request.session['message1']='您的帐号已激活,请登陆。'
            return redirect('/class/login/')
        else:
            request.session['message']='链接激活无效'
            return redirect('/class/register/')
    else:
        request.session['message'] ='链接激活无效或已过期'
        return redirect('/class/register/')

def error(request):
    return render(request,'404.html')

class Snowflake:
    def __init__(self, machine_id):
        self.machine_id = machine_id
        self.sequence = 0
        self.last_timestamp = -1

    def generate_id(self):
        timestamp = int(time.time() * 1000)
        if timestamp < self.last_timestamp:
            raise Exception("Clock moved backwards")
        if timestamp == self.last_timestamp:
            self.sequence = (self.sequence + 1) & 4095
            if self.sequence == 0:
                timestamp = self.wait_next_millis(self.last_timestamp)
        else:
            self.sequence = 0
        self.last_timestamp = timestamp
        return ((timestamp - 1288834974657) << 22) | (self.machine_id << 12) | self.sequence

    def wait_next_millis(self, last_timestamp):
        timestamp = int(time.time() * 1000)
        while timestamp <= last_timestamp:
            timestamp = int(time.time() * 1000)
        return timestamp



def task_uplaod(request):
    task_id=request.GET.get('task_id')
    request.session['task_id']=str(task_id)
    return render(request,'task_upload.html')

@login_required
def describe(request):
    staff=UserInfo.objects.get(username=request.user).is_staff
    task_id=request.GET.get('task_id')
    if task_id==None:
        task_id=request.POST.get('task_id')
    obj=task.objects.get(task_id=task_id)
    path=str(obj.task_file)
    task_type=obj.task_type
    task_name=obj.task_name
    start=obj.task_start
    end=obj.task_end
    status=obj.task_status
    if obj.describe=='4':
        task_describe='四人一组'
    elif obj.describe=='6':
        task_describe='六人一组'
    elif obj.describe=='8':
        task_describe='八人一组'
    elif obj.describe=='10':
        task_describe='十人一组'
    else:
        task_describe=obj.describe
    send=obj.send_name
    if status=='已截止' or status=='已完成':
        if task_type=='分组任务':
            file='/home/dreamcreator/class/media/xlsx/%s.xlsx'%task_id
            if os.path.exists(file):
                filepath=zip_ya(file)
                os.remove(file)
            elif os.path.exists(file+'.zip'):
                filepath=file+'.zip'
        else:
            file='/home/dreamcreator/class/media/'+str(task_id)
            file1=file+'.zip'
            if os.path.exists(file):
                filepath=zip_ya(file)
                shutil.rmtree(file, ignore_errors=False)
            elif os.path.exists(file1):
                filepath=file1
    if request.method=='GET':
        return render(request,'describe.html',locals())
    if request.method=='POST':
        file_name = os.path.basename(path)
        encoded_file_name = urllib.parse.quote(file_name.encode('utf-8'))
        file = open(path, 'rb')
        response = FileResponse(file, as_attachment=True, filename=encoded_file_name)
        return response


def send_zip(request):
    if request.method=='GET':
        task_id=request.GET.get('task_id')
        obj=task.objects.get(task_id=task_id)
        if obj.task_type=='分组任务':
            path='/home/dreamcreator/class/media/xlsx/%s.xlsx'%task_id+'.zip'
        else:
            path='/home/dreamcreator/class/media/'+str(task_id)+'.zip'
        file_name = os.path.basename(path)
        encoded_file_name = urllib.parse.quote(file_name.encode('utf-8'))
        if os.path.exists(path):
            file = open(path, 'rb')
            response = FileResponse(file, as_attachment=True, filename=encoded_file_name)
            return response
        else:
            return HttpResponse('没有数据')


def zip_ya(startdir):
    file_news = startdir +'.zip'
    z = zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED)
    for dirpath, dirnames, filenames in os.walk(startdir):
        fpath = dirpath.replace(startdir, '')
        fpath = fpath and fpath + os.sep or ''
        for filename in filenames:
            z.write(os.path.join(dirpath, filename), fpath + filename)
    z.close()
    return file_news

@login_required
def only(request):
    login_user = request.user
    user = UserInfo.objects.get(username=login_user)
    if request.method=='GET':
        stu_name=user.stu_name
        email=user.email
        stu_id=user.stu_id
        stu_work=user.stu_work
        stunum=user.stunum
    if request.method=='POST':
        if 'stu_name' in request.POST:
            stu_name=request.POST.get('stu_name')
            email=request.POST.get('email')
            if email:
                pan1 = re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email)
                if not pan1:
                    message1='请输入正确的邮箱!'
                    return render(request,'only.html',locals())
            if stu_name and not email:
                user.stu_name=stu_name
                message = '修改成功'
            elif not stu_name and email:
                send_email(email,user.id)
                message='邮箱修改激活链接已通过邮箱发送成功，请尽快激活！'
                request.session['email']=email
            else:
                user.stu_name=stu_name
                send_email(email, user.id)
                message = '邮箱修改激活链接已通过邮箱发送成功，请尽快激活！'
                request.session['email'] = email
            user.save()
        if 'password' in request.POST:
            oldpass=request.POST.get('oldpass')
            password=request.POST.get('password')
            repass=request.POST.get('repass')
            pan1=re.match(r'^(.+){6,12}$',password)
            if not pan1:
                message1='密码必须6到12位!'
                return render(request, 'only.html', locals())
            if password!=repass:
                message1='输入的两次密码不一致！'
            pan=authenticate(username=login_user,is_active=1,password=oldpass)
            if pan:
                send_pass(user.email,user.id)
                request.session['pass']=password
                message='密码修改链接已通过邮箱发送成功，请尽快激活!'
            else:
                message1='旧密码错误'
    return render(request,'only.html',locals())

@login_required
def add(request):
    login_user=request.user
    user=UserInfo.objects.get(username=login_user)
    if request.method=='GET':
        if user.stu_work in ['班长', '副班长', '学习委员', '学委']:
            return render(request,'add.html')
        else:
            message='你没有权限使用该功能！'
            return render(request,'admin-role.html',locals())
    if request.method=='POST':
        stu_name = request.POST.get('username')
        stu_id = request.POST.get('stu_id')
        stunum = request.POST.get('stunum')
        stu_work = request.POST.get('stu_work')
        uidb64 = urlsafe_base64_encode(force_bytes(stunum))
        pinyin = lazy_pinyin(stu_name)
        pinyin_abbr = ''.join([s[0] for s in pinyin])
        username = str(uidb64) + str(stu_id) + pinyin_abbr
        num = str(stu_id)[-4:]  # 学号后四位
        password = 'ZKSD' + pinyin_abbr + num
        obj = UserInfo(stu_name=stu_name, username=username, password=password, stunum=stunum, stu_id=stu_id,
                       class_num=user.class_num, stu_work=stu_work, is_active=1, stu_class=user.stu_class)
        if stu_work != '学生':
            obj.is_staff = 1
        obj.save()
        message1 = '添加成功!'
        return render(request,'add.html',locals())


def modify(request):
    uidb64 = request.GET.get('uid')
    token = request.GET.get('token')
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        # 获取对应的用户对象
        user = UserInfo.objects.get(id=uid)
    except (TypeError, ValueError, OverflowError, UserInfo.DoesNotExist):
        user = None
    conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=3)
    is_token = conn.get(user.id).decode('utf-8')
    if is_token:
        if user is not None and token == is_token:
            password=make_password(request.session.get('pass'))
            if password:
                del request.session['pass']
            user.password=password
            user.save()
            request.session['message1'] = '密码修改成功,请重新登录。'
            return redirect('/class/login/')
        else:
            request.session['message'] = '链接激活无效'
            return redirect('/class/register/')
    else:
        request.session['message'] = '链接激活无效或已过期'
        return redirect('/class/register/')

def changeemail(request):
    uidb64 = request.GET.get('uid')
    token = request.GET.get('token')
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        # 获取对应的用户对象
        user = UserInfo.objects.get(id=uid)
    except (TypeError, ValueError, OverflowError, UserInfo.DoesNotExist):
        user = None
    conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=3)
    is_token = conn.get(user.id).decode('utf-8')
    if is_token:
        if user is not None and token == is_token:
            email=request.session.get('email')
            if email:
                del request.session['email']
            user.email=email
            user.save()
            request.session['message1'] = '邮箱修改成功,请重新登录。'
            return redirect('/class/login/')
        else:
            request.session['message'] = '链接激活无效'
            return redirect('/class/register/')
    else:
        request.session['message'] = '链接激活无效或已过期'
        return redirect('/class/register/')


def changepass(request):
    uidb64 = request.GET.get('uid')
    token = request.GET.get('token')
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        # 获取对应的用户对象
        user = UserInfo.objects.get(id=uid)
    except (TypeError, ValueError, OverflowError, UserInfo.DoesNotExist):
        user = None
    conn = redis.StrictRedis(host="127.0.0.1", port=6379, password="", db=3)
    is_token = conn.get(user.id).decode('utf-8')
    if is_token:
        if user is not None and token == is_token:
            password=request.session.get('changepass')
            if password:
                del request.session['changepass']
            user.password=make_password(password)
            user.save()
            request.session['message1'] = '密码修改成功,请重新登录。'
            return redirect('/class/login/')
        else:
            request.session['message'] = '链接激活无效'
            return redirect('/class/register/')
    else:
        request.session['message'] = '链接激活无效或已过期'
        return redirect('/class/register/')

